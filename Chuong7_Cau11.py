#Câu 11: Xử lý Excel File - Viết phần mềm Quản Lý Nhân Viên
'''
Yêu cầu:
Viết phần mềm quản lý Nhân viên lưu bằng Excel. Mỗi nhân viên có Mã, Tên, Tuổi.
− Phần mềm cho phép lưu Nhân viên vào File Excel
− Phần mềm cho phép đọc danh sách Nhân viên trong File Excel
− Phần mềm cho phép sắp xếp Nhân viên theo Tuổi tăng dần
'''
from openpyxl import Workbook, load_workbook

# Đường dẫn file Excel
FILE_NAME = "NhanVien.xlsx"

# Hàm ghi danh sách nhân viên vào Excel
def ghi_nhanvien(danhsach):
    wb = Workbook()
    ws = wb.active
    ws.title = "NhanVien"
    ws.append(["STT", "Mã", "Tên", "Tuổi"])

    for i, nv in enumerate(danhsach, start=1):
        ws.append([i, nv["Mã"], nv["Tên"], nv["Tuổi"]])

    wb.save(FILE_NAME)
    print("✅ Đã lưu danh sách nhân viên vào file Excel!")

# Hàm đọc danh sách nhân viên từ Excel
def doc_nhanvien():
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    danhsach = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        danhsach.append({"STT": row[0], "Mã": row[1], "Tên": row[2], "Tuổi": row[3]})
    return danhsach

# Hàm sắp xếp theo tuổi tăng dần
def sapxep_theo_tuoi(danhsach):
    return sorted(danhsach, key=lambda nv: nv["Tuổi"])

# --- Ví dụ sử dụng ---
if __name__ == "__main__":
    # Bước 1: Lưu dữ liệu ban đầu
    ds = [
        {"Mã": "NV1", "Tên": "An", "Tuổi": 18},
        {"Mã": "NV2", "Tên": "Lành", "Tuổi": 22},
        {"Mã": "NV3", "Tên": "Giải", "Tuổi": 20},
        {"Mã": "NV4", "Tên": "Thoát", "Tuổi": 19},
        {"Mã": "NV5", "Tên": "Hạnh", "Tuổi": 25},
        {"Mã": "NV6", "Tên": "Phúc", "Tuổi": 24},
    ]
    ghi_nhanvien(ds)

    # Bước 2: Đọc danh sách từ Excel
    ds_doc = doc_nhanvien()
    print("\n📋 Danh sách đọc từ file:")
    for nv in ds_doc:
        print(nv)

    # Bước 3: Sắp xếp theo tuổi tăng dần
    ds_sapxep = sapxep_theo_tuoi(ds_doc)
    print("\n📈 Danh sách sau khi sắp xếp theo tuổi:")
    for nv in ds_sapxep:
        print(nv)
