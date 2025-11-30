#Câu 8: Xử lý đọc Excel File
'''
Yêu cầu:
Sử dụng thư viện openpyxl để đọc file excel ở câu trước.
'''
from openpyxl import load_workbook

# Mở file Excel
wb = load_workbook('demo.xlsx')

# In ra danh sách tên các sheet
print(wb.sheetnames)

# Lấy sheet đầu tiên
ws = wb[wb.sheetnames[0]]

# Duyệt qua từng hàng (row) trong sheet
for row in ws.values:
    for value in row:
        print(value, "\t", end='')
    print("")  # Xuống dòng sau mỗi hàng